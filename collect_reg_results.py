import os
from os.path import join
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
# Settings
project_dir = "C:/Users/yl10702/Shan Dropbox/Team Account/Health_Insurance_and_Worker_Health"
out_dir = join(project_dir, "Result/bootstrap")

# Process data
key = ['bootstrap-l1']
for keywords in key:
    df_tot = None
    for filename in os.listdir(out_dir):
        if keywords not in filename:
            continue
        df = pd.read_csv(join(out_dir, filename), delimiter = "\t", usecols=['Unnamed: 0', '(1)'])
        IV_name = filename.replace('result-reg-bootstrap-', '')
        IV_name = IV_name.replace('-FY_FE.txt', '')
        df.rename({'Unnamed: 0' : ' ', '(1)' : IV_name}, axis=1, inplace=True)
        if df_tot is None:
            df_tot = df
        else:
            df_tot = pd.concat([df_tot, df], axis=1)
        df_tot.to_excel(join(out_dir,"collect_"+keywords+".xlsx"), index=False)



# Combining multiple Excel sheets into one 
dest_wb = Workbook()

for root, dir, filenames in os.walk(out_dir):
    for file in filenames:
        if 'collect_' not in file:
            continue
        file_name = file.split('.')[0]
        file_name = file_name.replace('collect_', '')
        # Absolute Path for Excel files
        file_path = os.path.abspath(os.path.join(root, file))
        # Create new sheet in destination Workbook
        dest_wb.create_sheet(file_name)
        dest_ws = dest_wb[file_name]
        # Read source data
        source_wb = load_workbook(file_path)
        source_sheet = source_wb.active
        for row in source_sheet.rows:
            for cell in row:
                dest_ws[cell.coordinate] = cell.value

dest_wb.save(join(out_dir,'result-reg-all-bootstrap-second_stage-FY_FE.xlsx'))